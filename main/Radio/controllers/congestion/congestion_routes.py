import pandas as pd
from dateutil.parser import *
from flask import Blueprint, render_template, jsonify, request
from flask_login import login_required, current_user
from openpyxl import load_workbook
from main.Radio.controllers.api.radio_api_utils import subtract_days_from_date
from main.Radio.services.congestion_radio_service import CongestionRadioService
from main.utils.logging_config import configure_logging
from main.utils.utils import role_required

radio_congestion_bp = Blueprint(
    'radio_congestion_bp',
    __name__,
    static_folder='static',
    template_folder='templates'
)

ALLOWED_EXTENSIONS = {'xlsx'}


def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


@radio_congestion_bp.route("/congestion")
@login_required
@role_required(['USER_FH_RADIO', 'ADMIN', 'USER_RADIO'])
def show_congestion():
    df = CongestionRadioService.get_file_list()
    date_range = ['Du ' + subtract_days_from_date(row['EndDate'], 7) +
                  ' Au ' + row['EndDate'].strftime("%d-%m-%Y %H:%M")
                  for index, row in df.iterrows()]
    df['date_range'] = date_range
    return render_template("Radio/congestion/congestion.html",
                           df=df)


@radio_congestion_bp.route("/congestion/upload_file", methods=["POST"])
def upload_new_file():
    # Get the uploaded CSV file
    file = request.files['file']

    if file.filename == '':
        return jsonify({'status': 'error',
                        'message': 'Aucun fichier selectionné, Veuillez choisir un fichier puis réessayer !'})

    if file and allowed_file(file.filename):
        try:
            df = None
            wb = load_workbook(file, read_only=True)
            if 'export PRS' in wb.sheetnames:
                df = pd.read_excel(file,
                                   engine="openpyxl",
                                   sheet_name="export PRS",
                                   usecols=['Time', 'eNodeB Name', 'Integrity', 'VS.FEGE.RxMaxSpeed_Mbs(Mbps)'])
            else:
                df = pd.read_excel(file,
                                   engine="openpyxl",
                                   usecols=['Time', 'eNodeB Name', 'Integrity', 'VS.FEGE.RxMaxSpeed_Mbs(Mbps)'])

            df["EndDate"] = max(df.Time)
            end_date = max(df.Time)
            congestion_obj = CongestionRadioService.get_alarm_by_save_date(end_date)
            if congestion_obj:
                return jsonify({'status': 'error',
                                'message': 'Le fichier existe déjà, Veuillez choisir un autre fichier !'})
            else:
                CongestionRadioService.insert_new_file(df)

                logger = configure_logging()
                logger.info(f"Fichier congestion PRS {end_date} ajoute par {current_user.username}")

        except (Exception,):
            return jsonify({'status': 'error',
                            'message': "Il semblerait que le fichier n'est pas correct, Veuillez choisir un autre "
                                       "fichier!"})
    else:
        return jsonify({'status': 'error',
                        'message': 'Format du fichier non prise en compte, Veuillez choisir un autre fichier!'})
    return jsonify({'status': 'success',
                    'message': 'Fichier ajouté avec success !'})


@radio_congestion_bp.route("/congestion/delete", methods=['GET', 'POST'])
def delete_file():
    if request.method == 'POST':
        end_date = pd.to_datetime(parse(request.form['end_date'], fuzzy=True).strftime("%d-%m-%Y %H:%M"))
        CongestionRadioService.remove_file(end_date)

        logger = configure_logging()
        logger.info(f"Fichier congestion PRS {end_date} supprime par {current_user.username}")

    return jsonify({'status': 'success',
                    'message': 'Fichier supprimé avec success !'})
