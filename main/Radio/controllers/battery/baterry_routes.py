from datetime import datetime

import pandas as pd
from flask import Blueprint, render_template, request, jsonify
from flask_login import login_required, current_user
from dateutil.parser import *
from openpyxl import load_workbook

from main.Radio.services.battery_service import BatteryService
from main.utils.logging_config import configure_logging
from main.utils.utils import role_required

radio_battery_bp = Blueprint(
    'radio_battery_bp', __name__,
    static_folder='static',
    template_folder='templates'
)

ALLOWED_EXTENSIONS = {'xlsx'}


def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


@radio_battery_bp.route('/battery')
@login_required
@role_required(['USER_FH_RADIO', 'ADMIN', 'USER_FH'])
def battery_page():
    df = BatteryService.get_files_list()
    return render_template("Radio/battery/battery.html", df_battery=df)


@radio_battery_bp.route("/battery/delete", methods=['GET', 'POST'])
def delete_battery_file():
    if request.method == 'POST':
        creation_date = pd.to_datetime(parse(request.form['creation_date'], fuzzy=True).strftime("%d-%m-%Y %H:%M:%S"))
        BatteryService.remove_file(creation_date)

        logger = configure_logging()
        logger.info(f"Fichier PMON du {creation_date} supprime par {current_user.username}")

    return jsonify({'status': 'success',
                    'message': 'Fichier supprimé avec success !'})


@radio_battery_bp.route("/battery/upload_file", methods=["GET", "POST"])
def upload_new_file():
    # Get the uploaded CSV file
    file = request.files['battery_file']

    if file.filename == '':
        return jsonify({'status': 'error',
                        'message': 'Aucun fichier selectionné, Veuillez choisir un fichier puis réessayer !'})

    if file and allowed_file(file.filename):
        try:
            df = None
            wb = load_workbook(file, read_only=True)
            if 'Display Battery' in wb.sheetnames:
                df = pd.read_excel(file,
                                   engine="openpyxl",
                                   sheet_name="Display Battery",
                                   usecols=["NAME", "Remaining Capacity(%)", "Remaining Time(min)", "Power Cut Times"])
            else:
                df = pd.read_excel(file,
                                   engine="openpyxl",
                                   usecols=['NAME', 'Remaining Capacity(%)', 'Remaining Time(min)', 'Power Cut Times'])

            current_datetime = datetime.now()
            formatted_datetime = current_datetime.strftime("%Y-%m-%d %H:%M:%S")

            df["Creation Date"] = formatted_datetime
            df["Creation Date"] = pd.to_datetime(df["Creation Date"])

            BatteryService.add_new_file(df)
            logger = configure_logging()
            logger.info(f"Nouveau Fichier Battery ajoute par {current_user.username}")

        except (Exception,):
            return jsonify({'status': 'error',
                            'message': "Il semblerait que le fichier n'est pas correct, Veuillez choisir un autre "
                                       "fichier!"})
    else:
        return jsonify({'status': 'error',
                        'message': 'Format du fichier non prise en compte, Veuillez choisir un autre fichier!'})
    return jsonify({'status': 'success',
                    'message': 'Fichier ajouté avec success !'})
